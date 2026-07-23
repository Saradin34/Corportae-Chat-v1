"""Upload routes: chat attachments + avatars (user & group)."""
import csv
import html
import os
from io import BytesIO, StringIO

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy import desc, select
from sqlalchemy.ext.asyncio import AsyncSession

from ..config import settings
from ..database import get_db
from ..models import Chat, ChatMember, DownloadEvent, User
from ..permissions import get_effective_permissions
from ..schemas import (
    AvatarResult,
    DocumentPreviewOut,
    DocumentPreviewRequest,
    DownloadEventOut,
    DownloadLogRequest,
    UploadResult,
)
from ..security import get_current_user
from ..storage import save_avatar, save_upload
from ..ws_manager import manager

router = APIRouter(prefix="/api/uploads", tags=["uploads"])


async def _read_limited(file: UploadFile, max_mb: int) -> bytes:
    max_bytes = max_mb * 1024 * 1024
    data = await file.read(max_bytes + 1)
    if len(data) > max_bytes:
        raise HTTPException(status_code=413, detail=f"Файл превышает {max_mb} МБ")
    if not data:
        raise HTTPException(status_code=400, detail="Пустой файл")
    return data


@router.post("/file", response_model=UploadResult)
async def upload_file(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Upload an attachment (image or document). Returns metadata the client
    then sends with POST /chats/{id}/messages to create the message."""
    # group permission: must be allowed to send files OR images
    perms = await get_effective_permissions(db, user)
    if not (perms["can_send_files"] or perms["can_send_images"]):
        raise HTTPException(status_code=403, detail="Ваша группа не может отправлять вложения")
    # runtime-editable size limit (falls back to env default)
    from .. import app_settings
    max_mb = await app_settings.get(db, "max_upload_mb") or settings.MAX_UPLOAD_MB
    data = await _read_limited(file, int(max_mb))
    saved = save_upload(data, file.filename or "file", file.content_type)
    return UploadResult(
        kind=saved.kind,
        url=saved.url,
        thumb=saved.thumb_url,
        name=saved.name,
        size=saved.size,
        width=saved.width,
        height=saved.height,
    )


@router.post("/avatar", response_model=AvatarResult)
async def upload_avatar(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Upload the current user's avatar (square-cropped circle)."""
    data = await _read_limited(file, settings.MAX_AVATAR_MB)
    try:
        url = save_avatar(data, file.filename or "avatar.jpg")
    except Exception:
        raise HTTPException(status_code=400, detail="Не удалось обработать изображение")
    user.avatar_url = url
    await db.commit()
    await db.refresh(user)
    return AvatarResult(avatar_url=url)


@router.post("/chat/{chat_id}/avatar", response_model=AvatarResult)
async def upload_chat_avatar(
    chat_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Upload a group's avatar (group admins only)."""
    cm = (
        await db.execute(
            select(ChatMember).where(ChatMember.chat_id == chat_id, ChatMember.user_id == user.id)
        )
    ).scalar_one_or_none()
    if not cm:
        raise HTTPException(status_code=403, detail="Нет доступа к этому чату")
    chat = (await db.execute(select(Chat).where(Chat.id == chat_id))).scalar_one_or_none()
    if not chat:
        raise HTTPException(status_code=404, detail="Чат не найден")
    if chat.type == "private":
        raise HTTPException(status_code=400, detail="Нельзя менять аватар личного чата")
    if not cm.is_admin and user.role != "admin":
        raise HTTPException(status_code=403, detail="Только администратор группы может менять аватар")

    data = await _read_limited(file, settings.MAX_AVATAR_MB)
    try:
        url = save_avatar(data, file.filename or "avatar.jpg")
    except Exception:
        raise HTTPException(status_code=400, detail="Не удалось обработать изображение")
    chat.avatar_url = url
    await db.commit()

    members = list((await db.execute(select(ChatMember.user_id).where(ChatMember.chat_id == chat_id))).scalars().all())
    await manager.send_to_users(members, {"type": "chat_updated", "chat_id": chat_id})
    return AvatarResult(avatar_url=url)


# ---------- Document preview + download history ----------
def _uploads_path_from_url(url: str) -> str:
    if not url.startswith("/uploads/"):
        raise HTTPException(status_code=400, detail="Недопустимый адрес файла")
    rel = url[len("/uploads/"):].replace("\\", "/")
    # Prevent path traversal and limit to stored upload subfolders.
    if ".." in rel.split("/") or rel.startswith("/") or not rel:
        raise HTTPException(status_code=400, detail="Недопустимый адрес файла")
    path = os.path.abspath(os.path.join(settings.UPLOAD_DIR, rel))
    root = os.path.abspath(settings.UPLOAD_DIR)
    if not path.startswith(root + os.sep):
        raise HTTPException(status_code=400, detail="Недопустимый адрес файла")
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="Файл не найден")
    return path


def _ext_from(name: str, url: str) -> str:
    src = name or url
    return os.path.splitext(src.split("?", 1)[0])[1].lower().lstrip(".")


def _safe_preview_html(value: str) -> str:
    return html.escape(value or "", quote=True)


def _preview_text(path: str, name: str) -> DocumentPreviewOut:
    with open(path, "rb") as f:
        raw = f.read(700_000)
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError:
        text = raw.decode("cp1251", errors="replace")
    more = os.path.getsize(path) > len(raw)
    warn = ["Показаны первые ~700 КБ файла"] if more else []
    return DocumentPreviewOut(kind="html", name=name, html=f"<pre class='dv-text'>{_safe_preview_html(text)}</pre>", warnings=warn)


def _preview_csv(path: str, name: str) -> DocumentPreviewOut:
    with open(path, "rb") as f:
        raw = f.read(700_000)
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError:
        text = raw.decode("cp1251", errors="replace")
    rows = []
    reader = csv.reader(StringIO(text))
    for i, row in enumerate(reader):
        if i >= 80:
            break
        rows.append(row[:30])
    html_rows = "".join("<tr>" + "".join(f"<td>{_safe_preview_html(c)}</td>" for c in r) + "</tr>" for r in rows)
    return DocumentPreviewOut(kind="html", name=name, html=f"<div class='doc-table-wrap'><table class='doc-table'>{html_rows}</table></div>", warnings=[])


def _preview_docx(path: str, name: str) -> DocumentPreviewOut:
    try:
        import mammoth
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=503, detail=f"DOCX preview недоступен: {e}")
    with open(path, "rb") as f:
        result = mammoth.convert_to_html(f)
    warnings = [str(m.message) for m in getattr(result, "messages", [])[:5]]
    return DocumentPreviewOut(kind="html", name=name, html=f"<article class='docx-preview'>{result.value}</article>", warnings=warnings)


def _preview_xlsx(path: str, name: str) -> DocumentPreviewOut:
    try:
        from openpyxl import load_workbook
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=503, detail=f"XLSX preview недоступен: {e}")
    wb = load_workbook(path, read_only=True, data_only=True)
    parts = []
    warnings = []
    for si, ws in enumerate(wb.worksheets[:5]):
        parts.append(f"<h3>{_safe_preview_html(ws.title)}</h3><div class='doc-table-wrap'><table class='doc-table'>")
        for r_i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if r_i > 80:
                warnings.append(f"{ws.title}: показаны первые 80 строк")
                break
            cells = []
            for c in list(row)[:30]:
                cells.append(f"<td>{_safe_preview_html('' if c is None else str(c))}</td>")
            parts.append("<tr>" + "".join(cells) + "</tr>")
        parts.append("</table></div>")
    if len(wb.worksheets) > 5:
        warnings.append("Показаны первые 5 листов")
    return DocumentPreviewOut(kind="html", name=name, html="<div class='xlsx-preview'>" + "".join(parts) + "</div>", warnings=warnings)


@router.post("/preview", response_model=DocumentPreviewOut)
async def preview_document(
    data: DocumentPreviewRequest,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    path = _uploads_path_from_url(data.url)
    name = data.name or os.path.basename(path)
    ext = _ext_from(name, data.url)
    db.add(DownloadEvent(user_id=user.id, username=user.username, file_url=data.url, file_name=name, action="preview"))
    await db.commit()

    if ext == "pdf":
        return DocumentPreviewOut(kind="pdf", url=data.url, name=name)
    if ext in settings.IMAGE_EXTENSIONS.split(","):
        return DocumentPreviewOut(kind="image", url=data.url, name=name)
    if ext in ("docx",):
        return _preview_docx(path, name)
    if ext in ("xlsx", "xlsm"):
        return _preview_xlsx(path, name)
    if ext == "csv":
        return _preview_csv(path, name)
    if ext in ("txt", "log", "json", "xml", "md", "html", "htm"):
        return _preview_text(path, name)
    return DocumentPreviewOut(kind="unsupported", url=data.url, name=name, warnings=["Предпросмотр этого типа файла пока не поддерживается"])


@router.post("/download-log")
async def download_log(
    data: DownloadLogRequest,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    if data.action not in ("preview", "download", "open"):
        data.action = "download"
    # Validate own upload URL if possible; for robustness log even if file was removed.
    if not data.url.startswith("/uploads/"):
        raise HTTPException(status_code=400, detail="Недопустимый адрес файла")
    db.add(DownloadEvent(user_id=user.id, username=user.username, file_url=data.url, file_name=data.name, action=data.action))
    await db.commit()
    return {"ok": True}


@router.get("/download-history", response_model=list[DownloadEventOut])
async def download_history(
    url: str = "",
    limit: int = 30,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    stmt = select(DownloadEvent)
    if url:
        if not url.startswith("/uploads/"):
            raise HTTPException(status_code=400, detail="Недопустимый адрес файла")
        stmt = stmt.where(DownloadEvent.file_url == url)
    elif user.role != "admin":
        stmt = stmt.where(DownloadEvent.user_id == user.id)
    stmt = stmt.order_by(desc(DownloadEvent.id)).limit(max(1, min(limit, 200)))
    rows = (await db.execute(stmt)).scalars().all()
    return [DownloadEventOut.model_validate(r) for r in rows]
